from __future__ import annotations

import json
import gzip
import math
import re
import sys
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook
from docx import Document


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
from scripts.data_paths import external_data_root, path_label, resolve_source_path

DATA_ROOT = external_data_root()
INPUT = DATA_ROOT / "incoming" / "original_input_data_20260803"
UPDATED_INPUT = INPUT / "datos_actualizados"
DATA_SOURCE_CONFIG = ROOT / "config" / "data_sources.json"
DATA_SOURCES = json.loads(DATA_SOURCE_CONFIG.read_text(encoding="utf-8"))["sources"] if DATA_SOURCE_CONFIG.exists() else {}
LEGACY_DASHBOARD = ROOT / "apps" / "legacy-dashboard"
OUTPUT = LEGACY_DASHBOARD / "dashboard_data.js"
QUALITY_OUTPUT = LEGACY_DASHBOARD / "data_quality_report.json"
RECORDS_OUTPUT = LEGACY_DASHBOARD / "dashboard_records.json.gz"
RECORDS_SCRIPT_OUTPUT = LEGACY_DASHBOARD / "dashboard_records.js"
CAPTURE_GEOJSON_OUTPUT = LEGACY_DASHBOARD / "bizkaia_codigos_postales.geojson"
AW_AGGREGATES_OUTPUT = LEGACY_DASHBOARD / "aw_capture_aggregates.json.gz"
AW_MAX_ENTRY_KG = 50_000
AW_ANOMALIES_OUTPUT = DATA_ROOT / "processed" / "quality" / "aw_weight_anomalies.csv"
AW_ANOMALIES_DASHBOARD_OUTPUT = LEGACY_DASHBOARD / "aw_weight_anomalies.csv"
def config_path(domain: str, key: str, fallback: Path) -> Path:
    value = DATA_SOURCES.get(domain, {}).get(key)
    return resolve_source_path(value) if value else fallback


def config_paths(domain: str, key: str, fallback: list[Path]) -> list[Path]:
    values = DATA_SOURCES.get(domain, {}).get(key)
    return [resolve_source_path(value) for value in values] if values else fallback


AW_FAMILIES_INPUT = config_path("aw_families", "current", INPUT / "residuos_aw_familias.csv")
AW_EQUIVALENCES_INPUT = config_path("aw_equivalences", "current", INPUT / "residuos_salida_aw_equivalencias.csv")
AW_HISTORICAL_INPUT = config_path("captacion_aw", "historical", INPUT / "Registro detalles residuos_2026-2018.xlsx")
AW_SAMPLE_INPUT = config_path("captacion_aw", "sample", INPUT / "AST_AW_Ejemplo_Registro detalles residuos.xlsx")

PESADAS_MAIN_INPUT = config_path("pesadas", "historical", INPUT / "AST_Pesadas_Garbigunes_2023-YTD_enviar.ods")
PESADAS_UPDATED_INPUTS = config_paths("pesadas", "updates", [
    UPDATED_INPUT / "3. Transportes-pesadas" / "AST_Pesadas_Garbigunes_2025.ods",
    UPDATED_INPUT / "3. Transportes-pesadas" / "AST_Pesadas_Garbigunes_2026-YTD.ods",
])
INCIDENCIAS_MAIN_INPUT = config_path("incidencias", "historical", INPUT / "AST_2022-2026YTD_Incidencias_Vehículos.ods")
INCIDENCIAS_UPDATED_INPUT = config_paths("incidencias", "updates", [UPDATED_INPUT / "5. Flota" / "AST_2025-202606_Incidencias_Vehículos.ods"])[0]
FLOTA_INPUT = config_path("flota", "historical", INPUT / "AST_20260624_Flota.ods")
FLOTA_UPDATED_INPUT = config_path("flota", "preferred", UPDATED_INPUT / "5. Flota" / "AST_202606_Flota.ods")
GNC_INPUT = config_path("gnc", "historical", INPUT / "AST_20260624_Estaciones de servicio GNC.ods")
GNC_UPDATED_INPUT = config_path("gnc", "preferred", UPDATED_INPUT / "5. Flota" / "AST_20260630_Estaciones de servicio GNC.ods")
TALLERES_INPUT = config_path("talleres", "historical", INPUT / "AST_20260624_Talleres.ods")
TALLERES_UPDATED_INPUT = config_path("talleres", "preferred", UPDATED_INPUT / "5. Flota" / "AST_202606_Talleres.ods")
RUTAS_INPUT = config_path("rutas", "historical", INPUT / "AST_20260625_Rutas_transporte_garbigunes.ods")
RUTAS_UPDATED_INPUT = config_path("rutas", "preferred", UPDATED_INPUT / "6. Garbigunes rutas" / "AST_20260625_Rutas_transporte_garbigunes.ods")
REFUERZOS_MAIN_INPUT = config_path("refuerzos", "historical", INPUT / "AST_Refuerzos-TRANSPORTE GARBIGUNE.xlsx")
REFUERZOS_UPDATED_INPUTS = config_paths("refuerzos", "updates", [
    UPDATED_INPUT / "0. Refuerzos" / "Refuerzos-2025bis_Solo_GAR.xlsx",
    UPDATED_INPUT / "0. Refuerzos" / "Refuerzos-2026bis_solo_GAR.xlsx",
])
MOVIL_MAIN_INPUT = config_path("movil", "historical", INPUT / "AST_GarbigunesMovil-Todos.xlsx")
MOVIL_UPDATED_INPUT = config_paths("movil", "updates", [UPDATED_INPUT / "4. Transportes garbigune móvil" / "GarbigunesMovil-Todos.xlsx"])[0]
CONVENIOS_INPUT = config_path("convenios", "current", INPUT / "2026_Ayuntamientos_Conveniados_Garbigunes.ods")
GARBIKUNE_LOCATIONS_INPUT = config_path("garbigune_locations", "current", INPUT / "garbigunes_ubicaciones.csv")
SITE_ALIASES_INPUT = config_path("site_aliases", "current", INPUT / "site_aliases.csv")
CP_GEOJSON_INPUT = config_path("cp_geojson", "current", INPUT / "bizkaia_codigos_postales.geojson")
COBERTURAS_UPDATED_INPUT = config_path("coberturas", "current", UPDATED_INPUT / "1.Coberturas" / "Coberturas-solo_GAR.xlsx")
AW_UPDATED_INPUT = config_path("captacion_aw", "update_not_used", UPDATED_INPUT / "2. Entradas garbigunes" / "DetallesEntradasGarbiker.xlsx")
PERSONAL_HISTORICAL_INPUT = config_path(
    "personal_historico",
    "current",
    DATA_ROOT / "raw" / "historical" / "recursos" / "AST_Datos históricos personal Transporte Garbigune.docx",
)

ODS_NS = {
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}
TABLE_ATTR = "{urn:oasis:names:tc:opendocument:xmlns:table:1.0}"
OFFICE_ATTR = "{urn:oasis:names:tc:opendocument:xmlns:office:1.0}"


def clean_key(value: Any) -> str:
    return str(value).strip() if value is not None and not pd.isna(value) else ""


def cell_text(cell: ET.Element) -> str:
    values = ["".join(p.itertext()) for p in cell.findall(".//text:p", ODS_NS)]
    text = " ".join(value for value in values if value).strip()
    if text:
        return text

    for name in ("value", "date-value", "string-value", "time-value"):
        attr = OFFICE_ATTR + name
        if attr in cell.attrib:
            return str(cell.attrib[attr])
    return ""


def read_ods(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    with ZipFile(path) as archive:
        root = ET.fromstring(archive.read("content.xml"))

    tables = root.findall(".//table:table", ODS_NS)
    table = None
    if sheet_name:
        table = next((item for item in tables if item.attrib.get(TABLE_ATTR + "name") == sheet_name), None)
    else:
        table = tables[0] if tables else None
    if table is None:
        return pd.DataFrame()

    rows: list[list[str]] = []
    for row in table.findall("table:table-row", ODS_NS):
        repeat_rows = int(row.attrib.get(TABLE_ATTR + "number-rows-repeated", "1"))
        cells: list[str] = []
        for cell in list(row):
            if "table-cell" not in cell.tag:
                continue
            repeat_cols = int(cell.attrib.get(TABLE_ATTR + "number-columns-repeated", "1"))
            value = cell_text(cell)
            if repeat_cols > 100 and value == "":
                break
            cells.extend([value] * min(repeat_cols, 100))

        if any(clean_key(value) for value in cells):
            rows.extend([cells] * min(repeat_rows, 1))

    if not rows:
        return pd.DataFrame()

    width = max(len(row) for row in rows)
    header = rows[0] + [""] * (width - len(rows[0]))
    names: list[str] = []
    seen: Counter[str] = Counter()

    for index, name in enumerate(header):
        base = clean_key(name) or f"col_{index + 1}"
        seen[base] += 1
        names.append(base if seen[base] == 1 else f"{base}_{seen[base]}")

    data = [(row + [""] * (width - len(row)))[:width] for row in rows[1:]]
    frame = pd.DataFrame(data, columns=names)
    return frame.loc[:, [col for col in frame.columns if not col.startswith("col_") or frame[col].astype(str).str.strip().any()]]


def read_xlsx(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, engine="openpyxl")


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str)


def existing_paths(paths: list[Path]) -> list[Path]:
    return [path for path in paths if path.exists()]


def source_label(paths: list[Path]) -> str:
    return " + ".join(path_label(path) for path in paths)


def parse_dates(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", dayfirst=True)


def parse_number(series: pd.Series) -> pd.Series:
    fallback = pd.to_numeric(series, errors="coerce")
    text = series.astype(str).str.strip()
    has_comma = text.str.contains(",", regex=False)
    euro_text = text.str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
    euro_numeric = pd.to_numeric(euro_text, errors="coerce")
    plain_numeric = pd.to_numeric(text.str.replace(",", ".", regex=False), errors="coerce")
    numeric = fallback.where(fallback.notna(), plain_numeric)
    return numeric.where(~has_comma, euro_numeric).where(lambda values: values.notna(), fallback)


def combine_with_updated(main: pd.DataFrame, updated_frames: list[pd.DataFrame], key_columns: list[str]) -> pd.DataFrame:
    frames = [frame.copy() for frame in updated_frames if not frame.empty]
    if not frames:
        return main.copy()
    updated = pd.concat(frames, ignore_index=True)
    if main.empty:
        return updated

    main_clean = main.copy()
    key_frame = updated[key_columns].drop_duplicates()
    key_frame["_updated_source_row"] = True
    merged = main_clean.merge(key_frame, on=key_columns, how="left")
    main_keep = main_clean.loc[merged["_updated_source_row"].isna().to_numpy()]
    return pd.concat([main_keep, updated], ignore_index=True)


def combine_with_updated_date_windows(main: pd.DataFrame, updated_frames: list[pd.DataFrame], date_column: str) -> pd.DataFrame:
    frames = [frame.copy() for frame in updated_frames if not frame.empty]
    if not frames:
        return main.copy()
    if main.empty:
        return pd.concat(frames, ignore_index=True)

    main_clean = main.copy()
    main_dates = parse_dates(main_clean[date_column])
    keep = pd.Series(True, index=main_clean.index)
    for frame in frames:
        dates = parse_dates(frame[date_column])
        if dates.notna().any():
            keep &= ~(main_dates.ge(dates.min()) & main_dates.le(dates.max()))
    return pd.concat([main_clean.loc[keep], *frames], ignore_index=True)


def read_pesadas_sources() -> tuple[pd.DataFrame, list[Path]]:
    main = read_ods(PESADAS_MAIN_INPUT)
    updated_paths = existing_paths(PESADAS_UPDATED_INPUTS)
    updated_frames = [read_ods(path) for path in updated_paths]
    combined = combine_with_updated_date_windows(main, updated_frames, "Fecha")
    return combined, [PESADAS_MAIN_INPUT, *updated_paths]


def read_incidencias_sources() -> tuple[pd.DataFrame, list[Path]]:
    main = read_ods(INCIDENCIAS_MAIN_INPUT)
    updated_paths = [INCIDENCIAS_UPDATED_INPUT] if INCIDENCIAS_UPDATED_INPUT.exists() else []
    updated_frames = [read_ods(path) for path in updated_paths]
    combined = combine_with_updated_date_windows(main, updated_frames, "Fecha")
    return combined, [INCIDENCIAS_MAIN_INPUT, *updated_paths]


def read_refuerzos_sources() -> tuple[pd.DataFrame, list[Path]]:
    main = read_xlsx(REFUERZOS_MAIN_INPUT)
    updated_paths = existing_paths(REFUERZOS_UPDATED_INPUTS)
    updated_frames = [read_xlsx(path) for path in updated_paths]
    combined = combine_with_updated_date_windows(main, updated_frames, "Fecha")
    return combined, [REFUERZOS_MAIN_INPUT, *updated_paths]


def read_movil_sources() -> tuple[pd.DataFrame, list[Path]]:
    main = read_xlsx(MOVIL_MAIN_INPUT)
    updated_paths = [MOVIL_UPDATED_INPUT] if MOVIL_UPDATED_INPUT.exists() else []
    updated_frames = [read_xlsx(path) for path in updated_paths]
    combined = combine_with_updated_date_windows(main, updated_frames, "Fecha")
    return combined, [MOVIL_MAIN_INPUT, *updated_paths]


def preferred_existing(primary: Path, fallback: Path) -> Path:
    return primary if primary.exists() else fallback


def parse_number_value(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = clean_key(value)
    if not text:
        return None
    try:
        if "," in text:
            return float(text.replace(".", "").replace(",", "."))
        return float(text)
    except ValueError:
        try:
            return float(text.replace(",", "."))
        except ValueError:
            return None


def normalize_text(value: Any) -> str:
    text = clean_key(value).upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


SITE_ALIAS_CACHE: dict[str, dict[str, str]] | None = None


def read_site_aliases() -> dict[str, dict[str, str]]:
    global SITE_ALIAS_CACHE
    if SITE_ALIAS_CACHE is not None:
        return SITE_ALIAS_CACHE
    aliases: dict[str, dict[str, str]] = {}
    if SITE_ALIASES_INPUT.exists():
        frame = read_csv(SITE_ALIASES_INPUT)
        for _, row in frame.iterrows():
            active = clean_key(row.get("active")).lower()
            if active and active not in {"true", "1", "yes", "si", "s"}:
                continue
            raw_name = normalize_text(row.get("raw_name"))
            site_key = clean_key(row.get("site_key")).upper()
            if raw_name and site_key:
                aliases[raw_name] = {
                    "site_key": site_key,
                    "site_type": clean_key(row.get("site_type")) or "fixed",
                }
    SITE_ALIAS_CACHE = aliases
    return aliases


def normalize_site(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"^GARBIGUNE\s+", "", text).strip()
    alias = read_site_aliases().get(text)
    return alias["site_key"] if alias else text


def normalize_place(value: Any) -> dict[str, str]:
    text = normalize_text(value)
    alias = read_site_aliases().get(text)
    if alias:
        site_type = alias["site_type"]
        site_key = alias["site_key"] if site_type == "fixed" else ""
        return {"place_key": alias["site_key"], "place_type": site_type, "site_key": site_key}
    if text.startswith("GARBIGUNE "):
        site_key = normalize_site(text)
        return {"place_key": site_key, "place_type": "fixed", "site_key": site_key}
    return {"place_key": text, "place_type": "review", "site_key": ""}


def month_label(value: Any) -> str:
    if pd.isna(value):
        return ""
    dt = pd.Timestamp(value)
    return f"{dt.year}-{dt.month:02d}"


def format_cp(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    text = clean_key(value)
    if not text:
        return ""
    numeric = pd.to_numeric(pd.Series([text]), errors="coerce").iloc[0]
    if not pd.isna(numeric):
        return str(int(numeric)).zfill(5)
    digits = re.sub(r"\D+", "", text)
    return digits.zfill(5) if digits else ""


def parse_single_date(value: Any) -> pd.Timestamp | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return pd.Timestamp(value)
    if isinstance(value, date):
        return pd.Timestamp(value)
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return pd.Timestamp(parsed)


def records_from_series(series: pd.Series, label: str, value: str, limit: int | None = None) -> list[dict[str, Any]]:
    data = series.sort_values(ascending=False)
    if limit:
        data = data.head(limit)
    return [{label: clean_key(index), value: safe_num(amount)} for index, amount in data.items()]


def read_aw_family_table() -> tuple[dict[str, dict[str, str]], list[dict[str, str]]]:
    if not AW_FAMILIES_INPUT.exists():
        return {}, []
    table = pd.read_csv(AW_FAMILIES_INPUT).fillna("")
    required = {"residuo_aw", "familia_aw"}
    missing = required - set(table.columns)
    if missing:
        raise ValueError(f"{AW_FAMILIES_INPUT.name} missing columns: {', '.join(sorted(missing))}")
    table["residuo_key"] = table["residuo_aw"].map(clean_key).str.upper()
    table["familia_aw"] = table["familia_aw"].map(clean_key)
    table = table[table["residuo_key"].ne("") & table["familia_aw"].ne("")].copy()
    family_map = {
        row["residuo_key"]: {
            "family": row["familia_aw"],
            "subfamily": clean_key(row.get("subfamilia_aw", "")) or row["familia_aw"],
            "description": clean_key(row.get("descripcion_familia", "")),
            "examples": clean_key(row.get("ejemplos", "")),
            "criteria": clean_key(row.get("criterio", "")),
        }
        for _, row in table.iterrows()
    }
    legend_rows = []
    for family, group in table.groupby("familia_aw", sort=False):
        first = group.iloc[0]
        legend_rows.append(
            {
                "family": clean_key(family),
                "description": clean_key(first.get("descripcion_familia", "")),
                "examples": clean_key(first.get("ejemplos", "")),
                "criteria": clean_key(first.get("criterio", "")),
                "mappedWastes": int(group["residuo_key"].nunique()),
                "subfamilies": sorted(clean_key(value) for value in group.get("subfamilia_aw", pd.Series(dtype=str)).unique() if clean_key(value)),
            }
        )
    return family_map, legend_rows


def read_aw_equivalence_table() -> dict[str, list[str]]:
    if not AW_EQUIVALENCES_INPUT.exists():
        return {}
    table = pd.read_csv(AW_EQUIVALENCES_INPUT).fillna("")
    required = {"residuo_salida", "familias_aw"}
    missing = required - set(table.columns)
    if missing:
        raise ValueError(f"{AW_EQUIVALENCES_INPUT.name} missing columns: {', '.join(sorted(missing))}")
    bridge: dict[str, list[str]] = {}
    for _, row in table.iterrows():
        waste = clean_key(row.get("residuo_salida", "")).upper()
        families = [clean_key(item) for item in str(row.get("familias_aw", "")).split("|") if clean_key(item)]
        if waste and families:
            bridge[waste] = families
    return bridge


def driver_cluster_definitions() -> dict[str, dict[str, str]]:
    return {
        "high_productivity": {
            "label": "Alta productividad",
            "color": "#147d64",
            "description": "Muchos servicios/día y score ajustado alto.",
            "action": "Analizar su contexto operativo para identificar prácticas replicables.",
        },
        "high_load": {
            "label": "Carga pesada",
            "color": "#246fb2",
            "description": "Kg/servicio alto con actividad suficiente.",
            "action": "Distinguir desempeño de efecto mix de residuos/ruta antes de comparar.",
        },
        "intensive": {
            "label": "Alta actividad",
            "color": "#7a5aa6",
            "description": "Muchos servicios y días, perfil de cobertura intensiva.",
            "action": "Vigilar estabilidad y fatiga operativa; útil para dimensionamiento.",
        },
        "stable_regular": {
            "label": "Regular estable",
            "color": "#4f8b3b",
            "description": "Actividad comparable y variación diaria baja.",
            "action": "Perfil útil para planificación recurrente y sustituciones previsibles.",
        },
        "variable": {
            "label": "Variable",
            "color": "#c67b24",
            "description": "Actividad comparable, pero con variabilidad diaria alta.",
            "action": "Revisar calendario, rutas y asignaciones que explican la dispersión.",
        },
        "low_sample": {
            "label": "Muestra baja",
            "color": "#8a98a0",
            "description": "Pocos días o baja confianza estadística.",
            "action": "No usar para conclusiones individuales sin ampliar muestra.",
        },
        "review": {
            "label": "Revisar",
            "color": "#b8463f",
            "description": "Score bajo o señal atípica frente al contexto.",
            "action": "Contrastar datos, rutas y residuos antes de interpretar rendimiento.",
        },
    }


def readings_config() -> dict[str, Any]:
    return {
        "kindLabels": {
            "opportunity": "Oportunidad",
            "attention": "Atención",
            "context": "Contexto",
            "quality": "Calidad",
            "stable": "Estable",
        },
        "panelQuestions": {
            "summary": "¿Dónde hay señales operativas relevantes?",
            "sites": "¿Qué puntos, rutas y residuos explican la carga?",
            "capture": "¿De dónde vienen las entradas AW y qué flujos conviene revisar?",
            "fleet": "¿Qué vehículos o averías requieren atención?",
            "drivers": "¿Qué perfiles operativos existen y cuáles son comparables?",
            "resources": "¿Dónde hay presión de cobertura o problemas de datos?",
        },
        "methodNote": "Las lecturas automáticas se generan en navegador con configuración analítica preprocesada por Python y filtros activos del usuario.",
    }


def safe_num(value: Any, digits: int = 2) -> float:
    if value is None or pd.isna(value):
        return 0.0
    number = float(value)
    if math.isclose(number, round(number)):
        return float(round(number))
    return round(number, digits)


def percentage(numerator: float, denominator: float) -> float:
    return round((numerator / denominator) * 100, 1) if denominator else 0.0


def pct_change(current: float, previous: float) -> float | None:
    if not previous:
        return None
    return round(((current - previous) / previous) * 100, 1)


def detect_mobile_schema(frame: pd.DataFrame) -> dict[str, Any]:
    normalized = {col: normalize_text(col) for col in frame.columns}
    weight_cols = [col for col, name in normalized.items() if any(token in name for token in ("PESO", "KG", "KILO", "TON"))]
    waste_cols = [col for col, name in normalized.items() if any(token in name for token in ("RESIDUO", "FRACCION", "FRACC"))]
    return {
        "columns": [clean_key(col) for col in frame.columns],
        "hasWeight": bool(weight_cols),
        "hasWaste": bool(waste_cols),
        "weightColumns": weight_cols,
        "wasteColumns": waste_cols,
        "integrationMode": "weight_waste_candidate" if weight_cols and waste_cols else "movements_only",
        "decision": (
            "El archivo móvil contiene campos de peso y residuo; puede evaluarse su integración con salidas transportadas."
            if weight_cols and waste_cols
            else "El archivo móvil no contiene peso ni residuo; se mantiene separado como movimientos del servicio móvil."
        ),
    }


def build_mobile_resources(movil: pd.DataFrame) -> dict[str, Any]:
    schema = detect_mobile_schema(movil)
    frame = movil.copy()
    frame["origin"] = frame["Municipio Origen"].map(clean_key)
    frame["destination"] = frame["Garbigune Destino"].map(clean_key)
    frame["driver"] = frame["Conductor"].map(clean_key)
    frame["vehicle"] = frame["Matrícula"].map(clean_key)
    frame["month"] = frame["date"].map(month_label)
    return {
        "meta": {
            **schema,
            "rows": int(len(frame)),
            "from": str(frame["date"].min().date()) if len(frame) else "",
            "to": str(frame["date"].max().date()) if len(frame) else "",
        },
        "destinations": records_from_series(frame["destination"].value_counts(), "destination", "count", 9),
        "origins": records_from_series(frame["origin"].value_counts(), "origin", "count", 12),
        "byMonth": records_from_series(frame["month"].value_counts(), "month", "count"),
        "byDriver": records_from_series(frame["driver"].value_counts(), "driver", "count", 10),
        "byVehicle": records_from_series(frame["vehicle"].value_counts(), "vehicle", "count", 10),
        "originDestination": (
            frame.groupby(["origin", "destination"])
            .size()
            .reset_index(name="count")
            .sort_values("count", ascending=False)
            .head(15)
            .to_dict("records")
        ),
    }


def capture_source_path() -> Path:
    return AW_HISTORICAL_INPUT if AW_HISTORICAL_INPUT.exists() else AW_SAMPLE_INPUT


def read_aw_aggregate_records(source: Path, family_map: dict[str, dict[str, str]], location_keys: set[str]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    aggregates: dict[tuple[str, ...], dict[str, Any]] = {}
    total_rows = 0
    skipped_rows = 0
    anomalous_weight_rows = 0
    anomalous_weight_kg = 0.0
    anomalous_weight_examples: list[dict[str, Any]] = []
    anomalous_weight_records: list[dict[str, Any]] = []
    raw_columns: set[str] = set()
    first_date: pd.Timestamp | None = None
    last_date: pd.Timestamp | None = None

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            names = [clean_key(value) for value in header]
            raw_columns.update(name for name in names if name)
            index = {name: position for position, name in enumerate(names) if name}
            date_col = "Fecha y hora" if "Fecha y hora" in index else "Fecha"
            if date_col not in index:
                skipped_rows += max((sheet.max_row or 1) - 1, 0)
                continue

            for row_number, row in enumerate(rows, start=2):
                total_rows += 1
                get = lambda name: row[index[name]] if name in index and index[name] < len(row) else None
                parsed_date = parse_single_date(get(date_col))
                kg = parse_number_value(get("Peso")) if "Peso" in index else None
                if parsed_date is None or pd.isna(kg):
                    skipped_rows += 1
                    continue
                site = clean_key(get("Garbigune")).upper()
                site_key = normalize_site(site)
                waste = clean_key(get("Residuo")).upper()
                family = family_map.get(waste, {}).get("family", "SIN FAMILIA")
                subfamily = family_map.get(waste, {}).get("subfamily", "SIN SUBFAMILIA")
                cp = format_cp(get("C.P."))
                unit = clean_key(get("Unidad"))
                user_type = clean_key(get("Tipo Usuario"))
                origin_municipality = clean_key(get("Municipio origen"))
                account_municipality = clean_key(get("Municipio cuenta"))
                original_kg = float(kg)
                valid_kg = original_kg
                is_anomalous_weight = original_kg < 0 or original_kg > AW_MAX_ENTRY_KG
                if kg < 0 or kg > AW_MAX_ENTRY_KG:
                    anomalous_weight_rows += 1
                    anomalous_weight_kg += original_kg
                    valid_kg = 0.0
                    anomaly = {
                        "source_file": path_label(source),
                        "source_sheet": sheet.title,
                        "source_row": int(row_number),
                        "fecha": str(parsed_date.date()),
                        "garbigune": site,
                        "site_key": site_key,
                        "residuo_aw": waste,
                        "familia_aw": family,
                        "subfamilia_aw": subfamily,
                        "tipo_usuario": user_type,
                        "municipio_origen": origin_municipality,
                        "municipio_cuenta": account_municipality,
                        "cp": cp,
                        "unidad": unit,
                        "peso_original_kg": safe_num(original_kg, 2),
                        "peso_validado_kg": 0.0,
                        "umbral_kg": int(AW_MAX_ENTRY_KG),
                        "motivo": f"Peso individual negativo o superior a {AW_MAX_ENTRY_KG:,} kg",
                        "pregunta_cliente": "Confirmar si el peso es correcto o si corresponde a un error de captura, separador decimal/miles, unidad o medición.",
                        "accion_propuesta": "Mantener la entrada para conteos, excluir temporalmente el peso de toneladas hasta confirmación.",
                    }
                    anomalous_weight_records.append(anomaly)
                    if len(anomalous_weight_examples) < 10:
                        anomalous_weight_examples.append(
                            {
                                "date": anomaly["fecha"],
                                "site": site,
                                "waste": waste,
                                "cp": cp,
                                "unit": unit,
                                "kg": anomaly["peso_original_kg"],
                                "sourceSheet": sheet.title,
                                "sourceRow": int(row_number),
                            }
                        )
                if first_date is None or parsed_date < first_date:
                    first_date = parsed_date
                if last_date is None or parsed_date > last_date:
                    last_date = parsed_date

                month = month_label(parsed_date)
                key = (
                    month,
                    month,
                    site,
                    site_key,
                    waste,
                    family,
                    subfamily,
                    user_type,
                    cp,
                    unit,
                    "1" if site_key in location_keys else "0",
                )
                if key not in aggregates:
                    aggregates[key] = {
                        "date": f"{month}-01",
                        "month": key[1],
                        "site": site,
                        "site_key": site_key,
                        "waste": waste,
                        "waste_family": family,
                        "waste_subfamily": subfamily,
                        "user_type": key[7],
                        "origin_municipality": "",
                        "account_municipality": "",
                        "cp": cp,
                        "unit": key[9],
                        "site_has_location": key[10] == "1",
                        "kg": 0.0,
                        "rows": 0,
                        "entries": 0,
                        "anomalous_weight_rows": 0,
                        "anomalous_weight_kg": 0.0,
                        "_origin_counts": Counter(),
                        "_account_counts": Counter(),
                    }
                item = aggregates[key]
                item["kg"] += valid_kg
                item["rows"] += 1
                item["entries"] += 1
                if is_anomalous_weight:
                    item["anomalous_weight_rows"] += 1
                    item["anomalous_weight_kg"] += original_kg
                item["_origin_counts"][origin_municipality] += 1
                item["_account_counts"][account_municipality] += 1
    finally:
        workbook.close()

    records = list(aggregates.values())
    for item in records:
        item["kg"] = safe_num(item["kg"], 2)
        item["anomalous_weight_kg"] = safe_num(item["anomalous_weight_kg"], 2)
        origin_counts = item.pop("_origin_counts")
        account_counts = item.pop("_account_counts")
        item["origin_municipality"] = origin_counts.most_common(1)[0][0] if origin_counts else ""
        item["account_municipality"] = account_counts.most_common(1)[0][0] if account_counts else ""
    records.sort(key=lambda item: (item["date"], item["site"], item["cp"], item["waste"]))

    AW_ANOMALIES_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    AW_ANOMALIES_DASHBOARD_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    anomaly_frame = pd.DataFrame(anomalous_weight_records)
    anomaly_columns = [
        "source_file",
        "source_sheet",
        "source_row",
        "fecha",
        "garbigune",
        "site_key",
        "residuo_aw",
        "familia_aw",
        "subfamilia_aw",
        "tipo_usuario",
        "municipio_origen",
        "municipio_cuenta",
        "cp",
        "unidad",
        "peso_original_kg",
        "peso_validado_kg",
        "umbral_kg",
        "motivo",
        "pregunta_cliente",
        "accion_propuesta",
    ]
    if anomaly_frame.empty:
        anomaly_frame = pd.DataFrame(columns=anomaly_columns)
    else:
        anomaly_frame = anomaly_frame.loc[:, anomaly_columns].sort_values(["fecha", "garbigune", "residuo_aw", "source_row"])
    anomaly_frame.to_csv(AW_ANOMALIES_OUTPUT, index=False, encoding="utf-8-sig")
    anomaly_frame.to_csv(AW_ANOMALIES_DASHBOARD_OUTPUT, index=False, encoding="utf-8-sig")

    return records, {
        "rawRowsRead": int(total_rows),
        "skippedRows": int(skipped_rows),
        "anomalousWeightRows": int(anomalous_weight_rows),
        "anomalousWeightKg": safe_num(anomalous_weight_kg, 2),
        "anomalousWeightMaxKg": int(AW_MAX_ENTRY_KG),
        "anomalousWeightExamples": anomalous_weight_examples,
        "anomalousWeightFile": path_label(AW_ANOMALIES_OUTPUT),
        "anomalousWeightDashboardFile": AW_ANOMALIES_DASHBOARD_OUTPUT.name,
        "sourceColumns": sorted(raw_columns),
        "from": str(first_date.date()) if first_date is not None else "",
        "to": str(last_date.date()) if last_date is not None else "",
    }


def aggregate_count(records: pd.DataFrame, group_cols: list[str], output_cols: dict[str, str], kg_total: float) -> pd.DataFrame:
    grouped = records.groupby(group_cols, dropna=False).agg(kg=("kg", "sum"), entries=("entries", "sum"), rows=("rows", "sum"))
    for source_col, out_col in output_cols.items():
        grouped[out_col] = records.groupby(group_cols, dropna=False)[source_col].nunique()
    grouped = grouped.reset_index().sort_values("kg", ascending=False)
    grouped["tons"] = (grouped["kg"] / 1000).round(2)
    if kg_total:
        grouped["share"] = (grouped["kg"] / kg_total * 100).round(1)
    return grouped


def build_capture_data() -> dict[str, Any]:
    family_map, family_legend_base = read_aw_family_table()
    locations = read_csv(GARBIKUNE_LOCATIONS_INPUT)
    locations["site"] = locations["garbigune"].map(clean_key)
    locations["site_key"] = locations["site_key"].map(clean_key)
    locations["cp"] = locations["codigo_postal"].map(format_cp)
    locations["lat"] = pd.to_numeric(locations["lat"], errors="coerce")
    locations["lon"] = pd.to_numeric(locations["lon"], errors="coerce")

    geojson_path = CP_GEOJSON_INPUT
    geojson = json.loads(geojson_path.read_text(encoding="utf-8"))
    features = []
    geo_cps: set[str] = set()
    for feature in geojson.get("features", []):
        properties = feature.get("properties", {})
        cp = format_cp(properties.get("COD_POSTAL"))
        if not cp:
            continue
        geo_cps.add(cp)
        features.append(
            {
                "type": "Feature",
                "properties": {
                    "cp": cp,
                    "ine": clean_key(properties.get("CODIGO_INE")),
                },
                "geometry": feature.get("geometry"),
            }
        )
    CAPTURE_GEOJSON_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    CAPTURE_GEOJSON_OUTPUT.write_text(
        json.dumps({"type": "FeatureCollection", "features": features}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    source = capture_source_path()
    location_keys = set(locations["site_key"])
    records, source_meta = read_aw_aggregate_records(source, family_map, location_keys)
    frame = pd.DataFrame(records)
    if frame.empty:
        frame = pd.DataFrame(columns=["date", "month", "site", "site_key", "waste", "waste_family", "waste_subfamily", "user_type", "origin_municipality", "account_municipality", "cp", "kg", "unit", "rows", "entries", "site_has_location"])

    cp_known = frame["cp"].ne("")
    cp_geo_match = frame["cp"].isin(geo_cps)
    kg_total = float(frame["kg"].sum())
    kg_with_cp = float(frame.loc[cp_known, "kg"].sum())
    kg_with_geo = float(frame.loc[cp_geo_match, "kg"].sum())
    unmatched_cps = sorted(cp for cp in frame.loc[cp_known, "cp"].unique() if cp not in geo_cps)
    unmapped_wastes = sorted(waste for waste in frame["waste"].unique() if waste and waste not in family_map)
    unmapped_subfamilies = sorted(waste for waste in frame.loc[frame["waste_subfamily"].eq("SIN SUBFAMILIA"), "waste"].unique() if waste)
    by_cp = (
        frame.assign(cp_label=frame["cp"].where(frame["cp"].ne(""), "SIN CP"))
        .groupby("cp_label")
        .agg(kg=("kg", "sum"), entries=("entries", "sum"), rows=("rows", "sum"), sites=("site", "nunique"), wastes=("waste", "nunique"), waste_families=("waste_family", "nunique"), waste_subfamilies=("waste_subfamily", "nunique"))
        .reset_index()
        .rename(columns={"cp_label": "cp"})
        .sort_values("kg", ascending=False)
    )
    by_cp["tons"] = (by_cp["kg"] / 1000).round(2)

    by_site = (
        frame.groupby("site")
        .agg(kg=("kg", "sum"), entries=("entries", "sum"), rows=("rows", "sum"), cps=("cp", lambda values: int(values[values.ne("")].nunique())), wastes=("waste", "nunique"), waste_families=("waste_family", "nunique"), waste_subfamilies=("waste_subfamily", "nunique"))
        .reset_index()
        .sort_values("kg", ascending=False)
    )
    by_site["tons"] = (by_site["kg"] / 1000).round(2)
    by_family = (
        frame.groupby("waste_family")
        .agg(kg=("kg", "sum"), entries=("entries", "sum"), rows=("rows", "sum"), wastes=("waste", "nunique"))
        .reset_index()
        .rename(columns={"waste_family": "family"})
        .sort_values("kg", ascending=False)
    )
    by_family["tons"] = (by_family["kg"] / 1000).round(2)
    by_family["share"] = (by_family["kg"] / kg_total * 100).round(1) if kg_total else 0
    by_subfamily = (
        frame.groupby(["waste_family", "waste_subfamily"])
        .agg(kg=("kg", "sum"), entries=("entries", "sum"), rows=("rows", "sum"), wastes=("waste", "nunique"))
        .reset_index()
        .rename(columns={"waste_family": "family", "waste_subfamily": "subfamily"})
        .sort_values("kg", ascending=False)
    )
    by_subfamily["tons"] = (by_subfamily["kg"] / 1000).round(2)
    by_subfamily["share"] = (by_subfamily["kg"] / kg_total * 100).round(1) if kg_total else 0
    family_totals = by_family.set_index("family").to_dict("index")
    family_legend = []
    for item in family_legend_base:
        totals = family_totals.get(item["family"], {})
        family_legend.append(
            {
                **item,
                "tons": safe_num(totals.get("tons", 0), 2),
                "share": safe_num(totals.get("share", 0), 1),
                "activeWastes": int(totals.get("wastes", 0) or 0),
            }
        )
    known_legend_families = {item["family"] for item in family_legend_base}
    for family in sorted(set(by_family["family"]) - known_legend_families):
        totals = family_totals.get(family, {})
        family_legend.append(
            {
                "family": family,
                "description": "Residuos AW sin asignación específica en la tabla editable.",
                "examples": ", ".join(unmapped_wastes[:6]),
                "criteria": "Pendiente de clasificar en residuos_aw_familias.csv",
                "mappedWastes": 0,
                "tons": safe_num(totals.get("tons", 0), 2),
                "share": safe_num(totals.get("share", 0), 1),
                "activeWastes": int(totals.get("wastes", 0) or 0),
            }
        )
    family_legend = sorted(family_legend, key=lambda item: item.get("tons", 0), reverse=True)
    aggregate_payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": source.name,
        "entryMetricMethod": "entradas aditivas estimadas a nivel de agregado; el fichero histórico no contiene Entrada ID original, por lo que se usa el número de registros AW agregados como proxy estable para filtros.",
        "grain": ["month", "site", "cp", "waste", "family", "subfamily", "user_type"],
        "timeFilterNote": "Agregado mensual para mantener velocidad en navegador; los filtros diarios del panel global se aplican por mes completo en Captación AW.",
        "records": records,
    }
    with gzip.open(AW_AGGREGATES_OUTPUT, "wt", encoding="utf-8", compresslevel=9) as archive:
        json.dump(sanitize(aggregate_payload), archive, ensure_ascii=False, separators=(",", ":"))

    return {
        "meta": {
            "rows": int(frame["rows"].sum()),
            "aggregateRows": int(len(frame)),
            "entries": int(frame["entries"].sum()),
            "from": source_meta["from"],
            "to": source_meta["to"],
            "kg": safe_num(kg_total, 2),
            "tons": safe_num(kg_total / 1000, 2),
            "cpRows": int(frame.loc[cp_known, "rows"].sum()),
            "cpRowsShare": percentage(float(frame.loc[cp_known, "rows"].sum()), float(frame["rows"].sum())),
            "cpKgShare": percentage(kg_with_cp, kg_total),
            "geoKgShare": percentage(kg_with_geo, kg_total),
            "geoCpCount": int(len(geo_cps)),
            "unmatchedCps": unmatched_cps,
            "familySource": path_label(AW_FAMILIES_INPUT),
            "mappedWasteTypes": int(frame.loc[frame["waste"].isin(family_map.keys()), "waste"].nunique()),
            "unmappedWasteTypes": int(len(unmapped_wastes)),
            "unmappedWastes": unmapped_wastes,
            "unmappedSubfamilyTypes": int(len(unmapped_subfamilies)),
            "unmappedSubfamilies": unmapped_subfamilies,
            "source": path_label(source),
            "rawRowsRead": int(source_meta["rawRowsRead"]),
            "skippedRows": int(source_meta["skippedRows"]),
            "anomalousWeightRows": int(source_meta["anomalousWeightRows"]),
            "anomalousWeightKg": safe_num(source_meta["anomalousWeightKg"], 2),
            "anomalousWeightMaxKg": int(source_meta["anomalousWeightMaxKg"]),
            "anomalousWeightExamples": source_meta["anomalousWeightExamples"],
            "anomalousWeightFile": source_meta["anomalousWeightFile"],
            "anomalousWeightDashboardFile": source_meta["anomalousWeightDashboardFile"],
            "aggregateFile": AW_AGGREGATES_OUTPUT.name,
            "entryMetricMethod": aggregate_payload["entryMetricMethod"],
            "timeFilterNote": aggregate_payload["timeFilterNote"],
            "geoSource": path_label(CP_GEOJSON_INPUT),
            "geoFile": "bizkaia_codigos_postales.geojson",
            "locationSource": path_label(GARBIKUNE_LOCATIONS_INPUT),
            "scopeNote": "Histórico de entradas AW; no se suma a salidas transportadas.",
        },
        "records": [],
        "locations": locations[["site", "site_key", "cp", "direccion", "lat", "lon", "fuente"]].to_dict("records"),
        "cpGeojsonFile": "bizkaia_codigos_postales.geojson",
        "aggregateFile": AW_AGGREGATES_OUTPUT.name,
        "geoCps": sorted(geo_cps),
        "byCp": by_cp[["cp", "tons", "kg", "entries", "rows", "sites", "wastes", "waste_families", "waste_subfamilies"]].to_dict("records"),
        "bySite": by_site[["site", "tons", "kg", "entries", "rows", "cps", "wastes", "waste_families", "waste_subfamilies"]].to_dict("records"),
        "familyLegend": family_legend,
        "byFamily": by_family[["family", "tons", "kg", "entries", "rows", "wastes", "share"]].to_dict("records"),
        "bySubfamily": by_subfamily[["family", "subfamily", "tons", "kg", "entries", "rows", "wastes", "share"]].to_dict("records"),
        "waste": records_from_series(frame.groupby("waste")["kg"].sum(), "waste", "kg", 15),
        "userTypes": records_from_series(frame.groupby("user_type")["rows"].sum(), "user_type", "rows"),
    }


def read_personal_hours() -> dict[str, Any]:
    path = PERSONAL_HISTORICAL_INPUT
    document = Document(path)
    table = document.tables[0]
    rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
    headers = rows[0]
    entries = []
    for row in rows[1:]:
        if not row or row[0].upper() == "TOTAL":
            continue
        item = {"year": row[0]}
        for index, header in enumerate(headers[1:], start=1):
            item[header.lower()] = row[index]
        entries.append(item)
    return {"headers": headers, "rows": entries}


def extract_vehicle_plate(value: Any) -> str:
    match = re.search(r"([0-9]{4}-[A-Z]{3})", clean_key(value))
    return match.group(1) if match else clean_key(value)


def read_routes() -> dict[str, Any]:
    path = preferred_existing(RUTAS_UPDATED_INPUT, RUTAS_INPUT)
    vehicle_routes = read_ods(path, "Vehículos_-_Ruta")
    site_routes = read_ods(path, "Garbigune_-_Ruta")
    base_addresses = read_ods(path, "Base_-_Dirección")

    route_base_overrides = {
        "TARDE: AMOREBIETA-IGORRE": "AMOREBIETA",
    }
    route_base: dict[str, str] = {}
    route_vehicles: dict[str, set[str]] = {}
    vehicle_meta: dict[str, dict[str, Any]] = {}
    for _, row in vehicle_routes.iterrows():
        route = clean_key(row.get("GARBIGUNES")).upper()
        base = clean_key(row.get("BASE")).upper() or route_base_overrides.get(route, "")
        if not route or any(token in route for token in ("CAMIÓN", "GARBIGUNE MÓVIL", "SERVICIO")):
            continue
        if base:
            route_base[route] = base
        plates = [extract_vehicle_plate(row.get("Vehículo 1")), extract_vehicle_plate(row.get("Vehículo 2"))]
        for plate in plates:
            if not re.fullmatch(r"[0-9]{4}-[A-Z]{3}", plate):
                continue
            route_vehicles.setdefault(route, set()).add(plate)
            meta = vehicle_meta.setdefault(plate, {"assigned_routes": set(), "assigned_bases": set()})
            meta["assigned_routes"].add(route)
            if base:
                meta["assigned_bases"].add(base)

    site_meta: dict[str, dict[str, str]] = {}
    site_rows: list[dict[str, str]] = []
    for _, row in site_routes.iterrows():
        locality = clean_key(row.get("Localidad")).upper()
        site = normalize_site(locality)
        route = clean_key(row.get("RUTA 1")).upper()
        route2 = clean_key(row.get("RUTA 2")).upper()
        base = route_base.get(route, route_base_overrides.get(route, ""))
        if not site:
            continue
        site_meta[site] = {"site": site, "route": route, "route2": route2, "base": base, "locality": locality}
        site_rows.append({"site": site, "locality": locality, "route": route, "route2": route2, "base": base})

    return {
        "site_meta": site_meta,
        "vehicle_meta": {
            plate: {
                "vehicle": plate,
                "assigned_routes": " · ".join(sorted(meta["assigned_routes"])),
                "assigned_bases": " · ".join(sorted(meta["assigned_bases"])),
            }
            for plate, meta in vehicle_meta.items()
        },
        "base_addresses": base_addresses.rename(columns={"BASE": "base", "Dirección base": "address"}).to_dict("records"),
        "sites": site_rows,
        "base_overrides": [{"route": route, "base": base} for route, base in route_base_overrides.items()],
    }


def read_convenios() -> dict[str, Any]:
    path = CONVENIOS_INPUT
    sheets = {
        "firmado": read_ods(path, "Convenio_Firmado"),
        "sin_convenio": read_ods(path, "Sin_Convenio"),
        "sin_renovar": read_ods(path, "Convenio_Sin_Renovar"),
    }

    def normalize_convenio_frame(frame: pd.DataFrame) -> pd.DataFrame:
        if any("MUNICIPIO" in str(col).upper() for col in frame.columns):
            return frame
        header_index = None
        for index, row in frame.iterrows():
            if any("MUNICIPIO" in clean_key(value).upper() for value in row.values):
                header_index = index
                break
        if header_index is None:
            return frame
        header = [clean_key(value) or f"col_{idx + 1}" for idx, value in enumerate(frame.loc[header_index].tolist())]
        seen: Counter[str] = Counter()
        columns = []
        for idx, item in enumerate(header):
            seen[item] += 1
            columns.append(item if seen[item] == 1 else f"{item}_{seen[item]}")
        data = frame.iloc[header_index + 1 :].copy()
        data.columns = columns
        return data

    rows: list[dict[str, Any]] = []
    for status, frame in sheets.items():
        if frame.empty:
            continue
        frame = normalize_convenio_frame(frame)
        municipality_col = next((col for col in frame.columns if "MUNICIPIO" in col), "")
        population_col = next((col for col in frame.columns if "HAB" in col), "")
        signed_col = next((col for col in frame.columns if "FIRMA" in col), "")
        end_col = next((col for col in frame.columns if "FIN" in col), "")
        for _, row in frame.iterrows():
            municipality = clean_key(row.get(municipality_col)).upper()
            if not municipality or "UDALERRIA" in municipality or "CONVENIOS" in municipality:
                continue
            rows.append(
                {
                    "municipality": municipality,
                    "status": status,
                    "population": safe_num(parse_number(pd.Series([row.get(population_col)])).iloc[0], 0) if population_col else 0,
                    "signed": clean_key(row.get(signed_col)) if signed_col else "",
                    "ends": clean_key(row.get(end_col)) if end_col else "",
                }
            )

    status_labels = {"firmado": "Convenio firmado", "sin_convenio": "Sin convenio", "sin_renovar": "Sin renovar"}
    by_status = []
    for status, group in pd.DataFrame(rows).groupby("status") if rows else []:
        by_status.append(
            {
                "status": status_labels.get(status, status),
                "municipalities": int(len(group)),
                "population": safe_num(group["population"].sum(), 0),
            }
        )
    by_status.sort(key=lambda item: item["municipalities"], reverse=True)
    return {"rows": rows, "byStatus": by_status}


def build() -> dict[str, Any]:
    routes = read_routes()
    convenios = read_convenios()

    pesadas, pesadas_sources = read_pesadas_sources()
    pesadas["date"] = parse_dates(pesadas["Fecha"])
    pesadas["month"] = pesadas["date"].map(month_label)
    pesadas["weight_kg"] = parse_number(pesadas["PESO KG"])
    pesadas["vehicle"] = pesadas["Matrícula"].map(clean_key)
    pesadas["site"] = pesadas["Localización/Garbigume"].map(clean_key).str.upper()
    pesadas["site_key"] = pesadas["site"].map(normalize_site)
    pesadas["waste"] = pesadas["Residuo"].map(clean_key).str.upper()
    pesadas = pesadas[pesadas["date"].notna() & pesadas["weight_kg"].notna()]
    pesadas["base"] = pesadas["site_key"].map(lambda value: routes["site_meta"].get(value, {}).get("base", "SIN RUTA"))
    pesadas["route"] = pesadas["site_key"].map(lambda value: routes["site_meta"].get(value, {}).get("route", "SIN RUTA"))
    aw_equivalences = read_aw_equivalence_table()
    pesadas_wastes = sorted(clean_key(value).upper() for value in pesadas["waste"].dropna().unique() if clean_key(value))
    aw_equivalence_missing = [waste_name for waste_name in pesadas_wastes if waste_name not in aw_equivalences]

    capture = build_capture_data()

    flota_source = preferred_existing(FLOTA_UPDATED_INPUT, FLOTA_INPUT)
    flota = read_ods(flota_source)
    flota["plate"] = flota["Matrícula"].map(clean_key)
    flota["fuel"] = flota["COMBUSTIBLE"].map(clean_key)
    flota["registered"] = parse_dates(flota["Fecha matriculación"])
    flota["age_years"] = ((pd.Timestamp.today().normalize() - flota["registered"]).dt.days / 365.25).round(1)

    incidencias, incidencias_sources = read_incidencias_sources()
    incidencias = incidencias[incidencias["Area"].map(clean_key).str.upper().str.contains("GARBIGUNES", na=False)].copy()
    incidencias["date"] = parse_dates(incidencias["Fecha"])
    incidencias["month"] = incidencias["date"].map(month_label)
    incidencias["plate"] = incidencias["Vehículo/maquinaria"].map(extract_vehicle_plate)
    incidencias_records = incidencias[["date", "month", "plate", "Tipo avería", "subgrupo avería", "Proveedor"]].rename(
        columns={"Tipo avería": "type", "subgrupo avería": "subgroup", "Proveedor": "workshop"}
    )
    incidencias_records["date"] = incidencias_records["date"].dt.strftime("%Y-%m-%d")

    refuerzos, refuerzos_sources = read_refuerzos_sources()
    refuerzos["date"] = parse_dates(refuerzos["Fecha"])
    refuerzos["month"] = refuerzos["date"].map(month_label)
    refuerzos["place"] = refuerzos["Lugar"].map(clean_key)

    movil, movil_sources = read_movil_sources()
    movil["date"] = parse_dates(movil["Fecha"])
    movil["vehicle"] = movil["Matrícula"].map(clean_key)
    mobile_resources = build_mobile_resources(movil)

    by_month = pesadas.groupby("month").agg(weight_kg=("weight_kg", "sum"), trips=("weight_kg", "size")).reset_index()
    by_month["tons"] = (by_month["weight_kg"] / 1000).round(1)
    by_month["kg_per_trip"] = (by_month["weight_kg"] / by_month["trips"]).round(0)
    by_month = by_month.sort_values("month")
    last_month = by_month.iloc[-1].to_dict() if len(by_month) else {}
    previous_month = by_month.iloc[-2].to_dict() if len(by_month) > 1 else {}
    previous_year_month = {}
    if last_month:
        previous_year_label = f"{int(str(last_month['month'])[:4]) - 1}-{str(last_month['month'])[5:7]}"
        match = by_month[by_month["month"].eq(previous_year_label)]
        previous_year_month = match.iloc[0].to_dict() if len(match) else {}

    site = pesadas.groupby("site").agg(weight_kg=("weight_kg", "sum"), trips=("weight_kg", "size")).sort_values("weight_kg", ascending=False)
    site["tons"] = (site["weight_kg"] / 1000).round(1)
    site["kg_per_trip"] = (site["weight_kg"] / site["trips"]).round(0)

    waste = pesadas.groupby("waste").agg(weight_kg=("weight_kg", "sum"), trips=("weight_kg", "size")).sort_values("weight_kg", ascending=False)
    waste["tons"] = (waste["weight_kg"] / 1000).round(1)
    waste["share"] = (waste["weight_kg"] / waste["weight_kg"].sum() * 100).round(1)

    vehicle = pesadas.groupby("vehicle").agg(weight_kg=("weight_kg", "sum"), trips=("weight_kg", "size")).sort_values("weight_kg", ascending=False)
    vehicle["tons"] = (vehicle["weight_kg"] / 1000).round(1)
    vehicle["kg_per_trip"] = (vehicle["weight_kg"] / vehicle["trips"]).round(0)
    vehicle["incidents"] = incidencias.groupby("plate").size()
    vehicle["incidents"] = vehicle["incidents"].fillna(0).astype(int)
    vehicle = vehicle.join(flota.set_index("plate")[["fuel", "age_years", "Centro", "Marca"]], how="left")
    vehicle_route_meta = pd.DataFrame(routes["vehicle_meta"].values())
    if not vehicle_route_meta.empty:
        vehicle = vehicle.join(vehicle_route_meta.set_index("vehicle")[["assigned_routes", "assigned_bases"]], how="left")

    route = pesadas.groupby("route").agg(weight_kg=("weight_kg", "sum"), trips=("weight_kg", "size"), sites=("site", "nunique")).sort_values("weight_kg", ascending=False)
    route["tons"] = (route["weight_kg"] / 1000).round(1)
    route["kg_per_trip"] = (route["weight_kg"] / route["trips"]).round(0)
    base = pesadas.groupby("base").agg(weight_kg=("weight_kg", "sum"), trips=("weight_kg", "size"), sites=("site", "nunique"), routes=("route", "nunique")).sort_values("weight_kg", ascending=False)
    base["tons"] = (base["weight_kg"] / 1000).round(1)
    base["kg_per_trip"] = (base["weight_kg"] / base["trips"]).round(0)

    driver = pesadas.groupby("Conductor").agg(
        total_kg=("weight_kg", "sum"),
        total_services=("weight_kg", "size"),
        work_days=("date", lambda values: values.dt.date.nunique()),
        first_day=("date", "min"),
        last_day=("date", "max"),
        sites=("site", "nunique"),
        waste_types=("waste", "nunique"),
        vehicles=("vehicle", "nunique"),
    )
    driver["tons"] = (driver["total_kg"] / 1000).round(1)
    driver["kg_per_service"] = (driver["total_kg"] / driver["total_services"]).round(0)
    driver["services_per_day"] = (driver["total_services"] / driver["work_days"]).round(2)
    driver["tons_per_day"] = (driver["tons"] / driver["work_days"]).round(2)
    driver["active_span_days"] = (driver["last_day"] - driver["first_day"]).dt.days + 1
    driver["service_density"] = (driver["work_days"] / driver["active_span_days"] * 100).round(1)
    driver = driver.sort_values(["total_services", "total_kg"], ascending=False)

    driver_daily = pesadas.groupby(["Conductor", pesadas["date"].dt.date]).agg(
        services=("weight_kg", "size"),
        kg=("weight_kg", "sum"),
    )
    driver_daily_summary = driver_daily.groupby("Conductor").agg(
        max_services_day=("services", "max"),
        avg_daily_kg=("kg", "mean"),
        max_daily_kg=("kg", "max"),
    )
    driver = driver.join(driver_daily_summary, how="left")
    driver["avg_daily_kg"] = driver["avg_daily_kg"].round(0)
    driver["max_daily_kg"] = driver["max_daily_kg"].round(0)

    driver_waste = (
        pesadas.pivot_table(index="Conductor", columns="waste", values="weight_kg", aggfunc="sum", fill_value=0)
        .div(1000)
        .round(1)
    )
    top_driver_index = driver.head(12).index
    top_waste_index = waste.head(8).index

    ref_by_year = refuerzos.groupby(refuerzos["date"].dt.year).size().dropna()
    ref_by_month = refuerzos.groupby("month").size().reset_index(name="count")
    ref_by_place = refuerzos["place"].value_counts()

    total_weight = float(pesadas["weight_kg"].sum())
    total_trips = int(len(pesadas))
    total_ref = int(len(refuerzos.dropna(how="all")))
    total_inc = int(len(incidencias.dropna(how="all")))
    incidents_by_year = incidencias.groupby(incidencias["date"].dt.year).size().dropna()
    no_route = pesadas[pesadas["route"].eq("SIN RUTA")]
    no_base = pesadas[pesadas["base"].eq("SIN RUTA") | pesadas["base"].eq("")]
    missing_counts = {
        "site": int(pesadas["site"].eq("").sum()),
        "waste": int(pesadas["waste"].eq("").sum()),
        "vehicle": int(pesadas["vehicle"].eq("").sum()),
        "driver": int(pesadas["Conductor"].map(clean_key).eq("").sum()),
        "route": int(len(no_route)),
        "base": int(len(no_base)),
    }
    quality_checks = [
        {
            "check": "Rutas sin asignar",
            "status": "ok" if percentage(len(no_route), len(pesadas)) < 0.5 else "warning",
            "value": int(len(no_route)),
            "share": percentage(len(no_route), len(pesadas)),
            "detail": "Registros con route = SIN RUTA sobre pesadas operativas.",
        },
        {
            "check": "Bases sin asignar",
            "status": "ok" if percentage(len(no_base), len(pesadas)) < 0.5 else "warning",
            "value": int(len(no_base)),
            "share": percentage(len(no_base), len(pesadas)),
            "detail": "Registros sin base operativa tras aplicar reglas de normalización.",
        },
        {
            "check": "Campos críticos incompletos",
            "status": "ok" if sum(missing_counts[key] for key in ("site", "waste", "vehicle", "driver")) == 0 else "warning",
            "value": int(sum(missing_counts[key] for key in ("site", "waste", "vehicle", "driver"))),
            "share": percentage(sum(missing_counts[key] for key in ("site", "waste", "vehicle", "driver")), len(pesadas)),
            "detail": "Suma de vacíos en garbigune, residuo, vehículo y conductor.",
        },
        {
            "check": "Cobertura incidencias",
            "status": "ok" if incidencias["date"].min() <= pesadas["date"].min() else "warning",
            "value": total_inc,
            "share": 100.0,
            "detail": f"Incidencias Garbigunes disponibles de {incidencias['date'].min().date()} a {incidencias['date'].max().date()}.",
        },
        {
            "check": "Integración garbigune móvil",
            "status": "ok",
            "value": int(mobile_resources["meta"]["rows"]),
            "share": 100.0,
            "detail": mobile_resources["meta"]["decision"],
        },
        {
            "check": "Captación AW con CP",
            "status": "ok" if capture["meta"]["geoKgShare"] >= 50 else "warning",
            "value": int(capture["meta"]["rows"]),
            "share": capture["meta"]["geoKgShare"],
            "detail": f"{capture['meta']['geoKgShare']}% del peso AW cruza con polígonos de CP; {capture['meta']['cpKgShare']}% tiene CP informado.",
        },
        {
            "check": "Pesos AW plausibles",
            "status": "warning" if capture["meta"]["anomalousWeightRows"] else "ok",
            "value": int(capture["meta"]["anomalousWeightRows"]),
            "share": percentage(capture["meta"]["anomalousWeightRows"], capture["meta"]["rawRowsRead"]),
            "detail": f"Las entradas AW con peso negativo o superior a {capture['meta']['anomalousWeightMaxKg']:,} kg se mantienen en conteos, pero su peso queda en cuarentena hasta confirmación. CSV: {capture['meta']['anomalousWeightFile']}. Peso bruto en revisión: {capture['meta']['anomalousWeightKg']:,.0f} kg.",
        },
        {
            "check": "Residuos AW con familia",
            "status": "warning" if capture["meta"]["unmappedWasteTypes"] or capture["meta"]["unmappedSubfamilyTypes"] else "ok",
            "value": int(capture["meta"]["unmappedWasteTypes"] + capture["meta"]["unmappedSubfamilyTypes"]),
            "share": percentage(capture["meta"]["mappedWasteTypes"], capture["meta"]["mappedWasteTypes"] + capture["meta"]["unmappedWasteTypes"]),
            "detail": f"{capture['meta']['mappedWasteTypes']} residuos AW mapeados en {capture['meta']['familySource']}; sin familia: {', '.join(capture['meta']['unmappedWastes'][:6]) or 'ninguno'}; sin subfamilia: {', '.join(capture['meta']['unmappedSubfamilies'][:6]) or 'ninguno'}.",
        },
        {
            "check": "Equivalencias salida → AW",
            "status": "warning" if aw_equivalence_missing else "ok",
            "value": int(len(aw_equivalence_missing)),
            "share": percentage(len(pesadas_wastes) - len(aw_equivalence_missing), len(pesadas_wastes)),
            "detail": f"{len(pesadas_wastes) - len(aw_equivalence_missing)} de {len(pesadas_wastes)} residuos de salida tienen equivalencia AW en {AW_EQUIVALENCES_INPUT.name}; pendientes: {', '.join(aw_equivalence_missing) or 'ninguno'}.",
        },
    ]
    quality_status = "warning" if any(item["status"] == "warning" for item in quality_checks) else "ok"
    pesadas_records = pesadas[["date", "month", "site", "waste", "vehicle", "Conductor", "base", "route", "weight_kg"]].copy()
    pesadas_records["date"] = pesadas_records["date"].dt.strftime("%Y-%m-%d")
    pesadas_records = pesadas_records.rename(columns={"Conductor": "driver", "weight_kg": "kg"})

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFiles": sorted(
            path_label(path)
            for source_root in (DATA_ROOT / "raw", DATA_ROOT / "reference")
            if source_root.exists()
            for path in source_root.rglob("*")
            if path.is_file() and not path.name.startswith("~$") and path.name != ".DS_Store"
        ),
        "activeSources": {
            "policy": "Histórico principal y actualizaciones en la raíz externa de datos; en solapes se priorizan las actualizaciones.",
            "pesadas": source_label(pesadas_sources),
            "incidencias": source_label(incidencias_sources),
            "flota": path_label(flota_source),
            "rutas": path_label(preferred_existing(RUTAS_UPDATED_INPUT, RUTAS_INPUT)),
            "convenios": path_label(CONVENIOS_INPUT),
            "movil": source_label(movil_sources),
            "refuerzos": source_label(refuerzos_sources),
            "captacionAw": capture["meta"]["source"],
            "captacionAwPolicy": "Se mantiene la fuente histórica principal porque conserva C.P. y municipio origen; DetallesEntradasGarbiker.xlsx no sustituye esa información para el análisis de captación.",
            "familiasAw": path_label(AW_FAMILIES_INPUT),
            "equivalenciasAw": path_label(AW_EQUIVALENCES_INPUT),
            "garbiguneLocations": path_label(GARBIKUNE_LOCATIONS_INPUT),
            "cpGeojson": path_label(CP_GEOJSON_INPUT),
        },
        "records": {
            "pesadas": pesadas_records.to_dict("records"),
            "incidencias": incidencias_records.to_dict("records"),
            "aw": [],
        },
        "kpis": {
            "tons": safe_num(total_weight / 1000, 1),
            "trips": total_trips,
            "kgPerTrip": safe_num(total_weight / total_trips, 0),
            "activeVehicles": int(pesadas["vehicle"].nunique()),
            "drivers": int(pesadas["Conductor"].nunique()),
            "sites": int(pesadas["site"].nunique()),
            "routes": int(pesadas["route"].nunique()),
            "bases": int(pesadas["base"].nunique()),
            "wasteTypes": int(pesadas["waste"].nunique()),
            "refuerzos": total_ref,
            "incidents": total_inc,
            "incidentsPerVehicle": safe_num(total_inc / max(flota["plate"].nunique(), 1), 1),
        },
        "coverage": {
            "pesadasFrom": str(pesadas["date"].min().date()),
            "pesadasTo": str(pesadas["date"].max().date()),
            "incidenciasFrom": str(incidencias["date"].min().date()),
            "incidenciasTo": str(incidencias["date"].max().date()),
            "incidenciasSource": source_label(incidencias_sources),
            "flotaAsOf": "2026-06",
            "refuerzosFrom": str(refuerzos["date"].min().date()),
            "refuerzosTo": str(refuerzos["date"].max().date()),
            "movilFrom": str(movil["date"].min().date()),
            "movilTo": str(movil["date"].max().date()),
            "awFrom": capture["meta"]["from"],
            "awTo": capture["meta"]["to"],
            "routesAsOf": "2026-06-25",
            "conveniosAsOf": "2026",
        },
        "quality": {
            "status": quality_status,
            "checks": quality_checks,
            "missing": missing_counts,
            "routeMapping": {
                "baseOverrides": routes["base_overrides"],
                "noRouteRows": int(len(no_route)),
                "noRouteShare": percentage(len(no_route), len(pesadas)),
                "noBaseRows": int(len(no_base)),
                "noBaseShare": percentage(len(no_base), len(pesadas)),
                "noRouteSites": records_from_series(no_route["site"].map(clean_key).value_counts(), "site", "count", 12),
            }
        },
        "analytics": {
            "monthly": {
                "lastMonth": clean_key(last_month.get("month")),
                "momTons": pct_change(float(last_month.get("tons", 0)), float(previous_month.get("tons", 0))) if last_month and previous_month else None,
                "yoyTons": pct_change(float(last_month.get("tons", 0)), float(previous_year_month.get("tons", 0))) if last_month and previous_year_month else None,
                "momTrips": pct_change(float(last_month.get("trips", 0)), float(previous_month.get("trips", 0))) if last_month and previous_month else None,
                "yoyTrips": pct_change(float(last_month.get("trips", 0)), float(previous_year_month.get("trips", 0))) if last_month and previous_year_month else None,
            },
            "dominance": {
                "topSiteShare": percentage(float(site["weight_kg"].iloc[0]) if len(site) else 0, total_weight),
                "topWasteShare": percentage(float(waste["weight_kg"].iloc[0]) if len(waste) else 0, total_weight),
                "topRouteShare": percentage(float(route["weight_kg"].iloc[0]) if len(route) else 0, total_weight),
            },
            "precomputed": {
                "scope": "Sin filtros interactivos; base calculada en Python para reducir cálculo inicial en navegador.",
                "summaryTopSites": site[["tons", "trips", "kg_per_trip"]].head(12).reset_index().to_dict("records"),
                "summaryTopWaste": waste[["tons", "trips", "share"]].head(12).reset_index().to_dict("records"),
                "sitesWasteMatrix": pesadas.pivot_table(index="site", columns="waste", values="weight_kg", aggfunc="sum", fill_value=0)
                .div(1000)
                .round(1)
                .loc[site.head(10).index, waste.head(8).index]
                .reset_index()
                .to_dict("records"),
                "driverTopByServices": driver.head(12).reset_index().rename(columns={"Conductor": "driver"})[
                    ["driver", "total_services", "work_days", "services_per_day", "tons", "kg_per_service"]
                ].to_dict("records"),
                "driverTopByLoad": driver.sort_values("tons", ascending=False).head(12).reset_index().rename(columns={"Conductor": "driver"})[
                    ["driver", "tons", "total_services", "work_days", "kg_per_service", "tons_per_day"]
                ].to_dict("records"),
                "mobile": mobile_resources["meta"],
            },
        },
        "analyticsConfig": {
            "awWasteFamilyBridge": aw_equivalences,
            "awWasteFamilyBridgeSource": AW_EQUIVALENCES_INPUT.name,
            "driverClusters": driver_cluster_definitions(),
            "readings": readings_config(),
        },
        "summary": {
            "byMonth": by_month[["month", "tons", "trips", "kg_per_trip"]].to_dict("records"),
            "topSites": site[["tons", "trips", "kg_per_trip"]].head(12).reset_index().to_dict("records"),
            "topWaste": waste[["tons", "trips", "share"]].head(12).reset_index().to_dict("records"),
        },
        "sitesWaste": {
            "sites": site[["tons", "trips", "kg_per_trip"]].reset_index().to_dict("records"),
            "waste": waste[["tons", "trips", "share"]].reset_index().to_dict("records"),
            "routes": route[["tons", "trips", "kg_per_trip", "sites"]].reset_index().to_dict("records"),
            "bases": base[["tons", "trips", "kg_per_trip", "sites", "routes"]].reset_index().to_dict("records"),
            "matrix": pesadas.pivot_table(index="site", columns="waste", values="weight_kg", aggfunc="sum", fill_value=0)
            .div(1000)
            .round(1)
            .loc[site.head(10).index, waste.head(8).index]
            .reset_index()
            .to_dict("records"),
            "detailSample": {
                "rows": int(capture["meta"]["rows"]),
                "entries": int(capture["meta"]["entries"]),
                "citizensShare": percentage(sum(item["rows"] for item in capture["userTypes"] if clean_key(item["user_type"]).upper() == "CIUDADANO"), capture["meta"]["rows"]),
                "topWaste": capture["waste"][:8],
                "cpRowsShare": capture["meta"]["cpRowsShare"],
                "geoKgShare": capture["meta"]["geoKgShare"],
            },
        },
        "capture": capture,
        "fleet": {
            "vehicles": vehicle.reset_index().rename(columns={"index": "vehicle"}).to_dict("records"),
            "incidentsByYear": [{"year": str(int(index)), "count": int(value)} for index, value in incidents_by_year.items()],
            "fuelMix": records_from_series(flota["fuel"].value_counts(), "fuel", "count"),
            "incidentTypes": records_from_series(incidencias["Tipo avería"].map(clean_key).value_counts(), "type", "count"),
            "incidentSubgroups": records_from_series(incidencias["subgrupo avería"].map(clean_key).value_counts(), "subgroup", "count", 10),
            "workshops": records_from_series(incidencias["Proveedor"].map(clean_key).value_counts(), "workshop", "count", 8),
        },
        "drivers": {
            "drivers": driver.reset_index().rename(columns={"Conductor": "driver"}).to_dict("records"),
            "topByServices": driver.head(12).reset_index().rename(columns={"Conductor": "driver"})[
                ["driver", "total_services", "work_days", "services_per_day", "tons", "kg_per_service"]
            ].to_dict("records"),
            "topByLoad": driver.sort_values("tons", ascending=False).head(12).reset_index().rename(columns={"Conductor": "driver"})[
                ["driver", "tons", "total_services", "work_days", "kg_per_service", "tons_per_day"]
            ].to_dict("records"),
            "topByDailyProductivity": driver[driver["work_days"] >= 10]
            .sort_values("services_per_day", ascending=False)
            .head(12)
            .reset_index()
            .rename(columns={"Conductor": "driver"})[
                ["driver", "services_per_day", "tons_per_day", "work_days", "total_services", "kg_per_service"]
            ].to_dict("records"),
            "wasteComposition": driver_waste.loc[top_driver_index, top_waste_index].reset_index().rename(columns={"Conductor": "driver"}).to_dict("records"),
            "wasteMixByTopDriver": [
                {
                    "driver": clean_key(driver_name),
                    "waste": clean_key(waste_name),
                    "tons": safe_num(value, 1),
                }
                for driver_name, row in driver_waste.loc[top_driver_index, top_waste_index].iterrows()
                for waste_name, value in row.items()
                if value > 0
            ],
        },
        "resources": {
            "refuerzosByYear": [{"year": str(int(index)), "count": int(value)} for index, value in ref_by_year.items()],
            "refuerzosByMonth": ref_by_month.to_dict("records"),
            "refuerzosByPlace": records_from_series(ref_by_place, "place", "count", 10),
            "refuerzosByPerson": records_from_series(refuerzos["Cubierta por"].map(clean_key).value_counts(), "person", "count", 10),
            "mobile": mobile_resources,
            "mobileDestinations": mobile_resources["destinations"],
            "mobileOrigins": mobile_resources["origins"],
            "mobileByMonth": mobile_resources["byMonth"],
            "mobileByDriver": mobile_resources["byDriver"],
            "mobileByVehicle": mobile_resources["byVehicle"],
            "mobileOriginDestination": mobile_resources["originDestination"],
            "personalHours": read_personal_hours(),
            "convenios": convenios,
            "routes": {
                "sites": routes["sites"],
                "baseAddresses": routes["base_addresses"],
            },
        },
    }


def json_default(value: Any) -> Any:
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.isoformat()
    if pd.isna(value):
        return None
    return value


def sanitize(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: sanitize(item) for key, item in value.items()}
    if isinstance(value, list):
        return [sanitize(item) for item in value]
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if pd.isna(value):
        return None
    return value


if __name__ == "__main__":
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    payload = sanitize(build())
    core_records = payload.get("records", {})
    payload["recordsFile"] = RECORDS_OUTPUT.name
    payload["recordsScriptFile"] = RECORDS_SCRIPT_OUTPUT.name
    payload["records"] = {"pesadas": [], "incidencias": [], "aw": []}
    with gzip.open(RECORDS_OUTPUT, "wt", encoding="utf-8", compresslevel=9) as archive:
        json.dump({"generatedAt": payload["generatedAt"], "records": core_records}, archive, ensure_ascii=False, separators=(",", ":"))
    RECORDS_SCRIPT_OUTPUT.write_text(
        "window.DASHBOARD_RECORDS = "
        + json.dumps({"generatedAt": payload["generatedAt"], "records": core_records}, ensure_ascii=False, separators=(",", ":"), default=json_default)
        + ";\n",
        encoding="utf-8",
    )
    content = "window.DASHBOARD_DATA = "
    content += json.dumps(payload, ensure_ascii=False, indent=2, default=json_default)
    content += ";\n"
    OUTPUT.write_text(content, encoding="utf-8")
    QUALITY_OUTPUT.write_text(
        json.dumps(
            {
                "generatedAt": payload["generatedAt"],
                "sourceFiles": payload["sourceFiles"],
                "activeSources": payload["activeSources"],
                "coverage": payload["coverage"],
                "quality": payload["quality"],
                "analytics": payload["analytics"],
            },
            ensure_ascii=False,
            indent=2,
            default=json_default,
        )
        + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT}")
    print(f"Wrote {RECORDS_OUTPUT}")
    print(f"Wrote {RECORDS_SCRIPT_OUTPUT}")
    print(f"Wrote {QUALITY_OUTPUT}")
